from django.db import models
from django.utils import timezone
from django.db.models.signals import post_save, pre_save
from django.dispatch import receiver
from openpyxl import load_workbook
import csv
import re
from .storage import OverwriteStorage
from datetime import datetime, time
from .controller import ControllerDB

TYPE_EXTENSION_FILE = (
    ('csv', 'csv'),
    ('xlsx', 'xlsx')
)


class File(models.Model):
    """This class will be storing data of file."""

    file = models.FileField(blank=False, null=False, storage=OverwriteStorage())
    type_file = models.CharField(
        max_length=4,
        choices=TYPE_EXTENSION_FILE,
    )
    table_name = models.CharField(max_length=120, verbose_name='Название файла')
    date_created = models.DateTimeField(verbose_name='Дата загрузки', default=timezone.now)
    date_updated = models.DateTimeField(verbose_name='Дата обновления', default=None, null=True)

    def __str__(self):
        return self.table_name

    @staticmethod
    def compare_previous_column():
        pass

    @staticmethod
    def get_type_cell_from_xlxs_file(cell):
        if type(cell) == str:
            if File.str_get_type_value(cell) in ('text', 'varchar'):
                if len(cell) > 128:
                    return 'text'
                else:
                    return 'varchar'
            else:
                return File.str_get_type_value(cell)
        elif type(cell) == int:
            return 'int'
        elif type(cell) == float:
            return 'float'
        elif isinstance(cell, datetime) == True:
            return 'datetime'
        elif isinstance(cell, time) == True:
            return 'time'
        elif cell is None:
            return ''
        else:
            print(f'cell.is_date() - {type(cell)}')
            return 'None1'

    @staticmethod
    def str_get_type_value(row):
        matching = re.findall(r'\d+', row)
        # checking if row is type of number
        if len(matching) == 1:
            return 'int'
        elif len(matching) >= 2:
            matching = re.findall(r'^\d*.\d*$', row)
            result = matching
            if len(matching) == 1:
                return 'float'

        # checking if type row is string
        matching = re.findall(r'\w+', row)
        if len(matching) == 1:
            if len(matching[0]) > 128:
                return 'text'
            else:
                return 'varchar'

        # checking if type row is datetime
        matching = re.findall(r'^\d{2}.\d{2}.\d{4} \d{2}.\d{2}.\d{2}$', row)
        if len(matching) == 1:
            return 'datetime'
        elif len(matching) == 0:
            matching = re.findall(r'\d{2}.\d{2}.\d{4}', row)
            if len(matching) == 1:
                return 'datetime'
            else:
                matching = re.findall(r'\d{4}.\d{2}.\d{2}', row)
                if len(matching) == 1:
                    return 'datetime'

        # checking if type row is time
        matching = re.findall(r'^\d{2}.\d{2}.\d{2}$', row)
        if len(matching) == 1:
            return 'time'

        # otherwise will be return empty row
        return ''

    @staticmethod
    def read_xlsx_file(file_path):
        print(f'file_path - {file_path}')
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        # get max row count
        max_row = sheet.max_row

        # get max column count
        max_column = sheet.max_column

        # iterate over all cells
        # iterate over all rows
        field_name_list = []
        type_fields_dict = {}
        rows_list = []
        for i in range(1, max_row + 1):
            # iterate over all columns
            midl_dict = {}
            for j in range(1, max_column + 1):
                # get particular cell value
                cell_obj = sheet.cell(row=i, column=j)

                # pass first row in excel file for get all value
                if i > 1:
                    if i > 2:
                        # compare current row coloumn type with previous
                        if type_fields_dict[j] != File.get_type_cell_from_xlxs_file(cell_obj.value):
                            # if that column type is empty that get current column
                            if type_fields_dict[j] == '':
                                type_fields_dict[j] = File.get_type_cell_from_xlxs_file(cell_obj.value)
                                # if current column type is empty do nothing, otherwise set type of 'str
                            elif File.get_type_cell_from_xlxs_file(cell_obj.value) == '':
                                pass
                            else:
                                type_fields_dict[j] = 'text'
                    else:
                        type_fields_dict[j] = File.get_type_cell_from_xlxs_file(cell_obj.value)

                    type_cell = File.get_type_cell_from_xlxs_file(cell_obj.value)
                    # if column is None that write  empty( '' ) value
                    if type_cell == '':
                        midl_dict[field_name_list[j - 1]] = 'NULL'
                    else:
                        midl_dict[field_name_list[j - 1]] = cell_obj.value

                else:
                    # get fields name
                    type_cell = ''
                    field_name_list.append(cell_obj.value)
                # print cell value

                # print(cell_obj.value, type_cell, end=' | ')
            if i > 1:
                rows_list.append(midl_dict)
            # print new line
            # print('\n')
        # print('----------' * 7)
        # print(rows_list)
        # print('----------' * 7)
        # print(field_name_list)
        # print('----------' * 7)
        # print(type_fields_dict)
        return field_name_list, type_fields_dict, rows_list

    @staticmethod
    def read_csv_file(file_path):
        def get_type_row_in_csv(row, type_fields_dict):
            """This method compare types fields for every coloums.
            If current type field is equal previous do nothing.
            And if it is false set type of 'text'"""

            for index, item in enumerate(row):
                # previous type is empty
                if type_fields_dict[index + 1] == '':
                    # if current type is empty too
                    if File.str_get_type_value(item) == '':
                        # set type of 'text'
                        type_fields_dict[index + 1] = 'text'
                    else:
                        # else set current type
                        type_fields_dict[index + 1] = File.str_get_type_value(item)
                # compate previous type field with current
                elif type_fields_dict[index + 1] != File.str_get_type_value(item):
                    # set type of 'text'
                    type_fields_dict[index + 1] = 'text'
                # otherwise set type of 'text'
                elif File.str_get_type_value(item) == '':
                    type_fields_dict[index + 1] = 'text'

            return type_fields_dict

        with open(file_path, 'r') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            field_name_list = []
            type_fields_dict = {}
            rows_list = []
            for number_line, row in enumerate(csv_reader):
                if number_line == 0:
                    # print(f'Column names are {", ".join(row)}')
                    for index in range(len(row)):
                        type_fields_dict[index + 1] = ''
                    field_name_list = row
                else:
                    dict_row = {}
                    type_fields_dict = get_type_row_in_csv(row, type_fields_dict)
                    for index, coloumn in enumerate(row):
                        if coloumn is None:
                            coloumn = 'NULL'
                        dict_row[field_name_list[index]] = coloumn
                    rows_list.append(dict_row)

            # print(f'row_list - {rows_list}')
            # print(f'field_name_list - {field_name_list}')
            # print(f'type_fields_dict - {type_fields_dict}')

            return field_name_list, type_fields_dict, rows_list

@receiver(post_save, sender=File)
def create_table(instance, **kwargs):
    # print(f'table_name - {instance.table_name}')
    if instance.type_file == 'xlsx':
        field_name_list, type_fields_dict, rows_list = File.read_xlsx_file(instance.file.path)
        controller = ControllerDB(
            table_name=instance.table_name,
            field_name_list=field_name_list,
            type_fields_dict=type_fields_dict,
            rows_list=rows_list,
        )
        controller.run_xlsx()
    elif instance.type_file == 'csv':
        field_name_list, type_fields_dict, rows_list = File.read_csv_file(instance.file.path)
        controller = ControllerDB(
            table_name=instance.table_name,
            field_name_list=field_name_list,
            type_fields_dict=type_fields_dict,
            rows_list=rows_list,
        )
        controller.run_csv()
        # print('here will be desion for .scv')
